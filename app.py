"""
iRAD PDF → Excel Report Generator
NO AI VERSION — uses PyMuPDF + regex to parse the structured iRAD PDF format

Requirements:
    pip install streamlit pymupdf openpyxl

Usage:
    py -m streamlit run irad_app_excel_only.py
"""

import io
import re
import traceback
from datetime import datetime

import streamlit as st

st.set_page_config(
    page_title="iRAD Excel Generator",
    page_icon="🚨",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Mono', monospace; }
.stApp { background: linear-gradient(135deg, #1A1A2E 0%, #16213E 50%, #0D1B2A 100%); }
.main-header {
    background: linear-gradient(90deg, #1A1A2E, #0F3460);
    border-bottom: 2px solid #E63946;
    padding: 18px 28px; margin: -1rem -1rem 2rem -1rem;
}
.badge {
    background:#E63946; color:white; font-weight:700; font-size:11px;
    padding:3px 10px; border-radius:3px; letter-spacing:2px;
}
.section-label {
    font-size:11px; font-weight:700; color:#E63946; letter-spacing:2px;
    text-transform:uppercase; border-bottom:1px solid rgba(230,57,70,0.25);
    padding-bottom:8px; margin-bottom:14px;
}
.step-row {
    display:flex; align-items:center; gap:10px; padding:8px 0;
    border-bottom:1px solid rgba(255,255,255,0.04); font-size:13px;
}
.stButton > button {
    background:#E63946; color:white; border:none; border-radius:6px;
    font-family:'IBM Plex Mono',monospace; font-weight:700;
    letter-spacing:1px; text-transform:uppercase; padding:10px 24px;
}
.stButton > button:hover { background:#c0313d; }
.stDownloadButton > button {
    background: linear-gradient(135deg, #10B981, #059669);
    color:white; border:none; border-radius:6px;
    font-family:'IBM Plex Mono',monospace; font-weight:700;
    letter-spacing:1px; text-transform:uppercase;
}
.log-box {
    background:rgba(0,0,0,0.35); border:1px solid rgba(255,255,255,0.06);
    border-radius:8px; padding:12px 16px; font-size:11px; color:#94A3B8;
    max-height:200px; overflow-y:auto; line-height:1.8;
}
.log-time { color:#E63946; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <span class="badge">iRAD</span>
  <span style="color:white;font-size:20px;font-weight:700;margin-left:12px">
    Accident Excel Generator
  </span>
  <span style="color:#94A3B8;font-size:12px;margin-left:12px">
    PDF → Excel &nbsp;|&nbsp; No AI Required
  </span>
</div>
""", unsafe_allow_html=True)

# ── session state ─────────────────────────────────────────────────────────────
if "logs" not in st.session_state:
    st.session_state.logs = []


# ══════════════════════════════════════════════════════════════════════════════
#  PDF PARSING  (PyMuPDF + regex — no AI)
# ══════════════════════════════════════════════════════════════════════════════

def extract_text_from_pdf(file_bytes: bytes) -> str:
    import fitz
    doc  = fitz.open(stream=file_bytes, filetype="pdf")
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    return text


def get(text: str, *patterns) -> str:
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
        if m:
            val = m.group(1).strip()
            val = re.split(r'\n[A-Z][a-z]', val)[0].strip()
            return val
    return ""


def get_int(text: str, *patterns) -> int:
    v = get(text, *patterns)
    try:
        return int(v)
    except Exception:
        return 0


def parse_irad_pdf(text: str) -> dict:
    d = {}

    d["firNumber"]           = get(text, r"FIR/CSR Number\s+([\w/]+)")
    d["firDate"]             = get(text, r"FIR Date & Time\s+([\d\-A-Za-z]+\s*:\s*[\d:]+\s*[APM]+)")
    d["stationName"]         = get(text, r"Station Name\s+([A-Z][A-Z ]+?)(?:\s{2,}|\n)")
    d["investigatingOfficer"]= get(text, r"Investigating(?:g)? Officer\s+([A-Za-z ]+?)(?:\n|Field Officer)")
    d["fieldOfficer"]        = get(text, r"Field Officer\s+([A-Za-z ]+?)(?:\n|District)")
    d["districtCode"]        = get(text, r"District code\s+(\d+)")
    d["districtName"]        = get(text, r"District Name\s+([A-Z][A-Z ]+?)(?:\s{2,}|\n|Act)")

    act_block = get(text, r"Act\s+(Bharatiya.*?)\s*Section", r"Act\s+([\w ,\d]+)\n")
    d["act"]      = act_block.replace("\n", " ").strip()
    d["sections"] = get(text, r"Section\s+(.*?)\s*State Rule",
                         r"(Section[\d\w(),\n ]+?)(?:State Rule|Accident Details)")

    d["accidentDate"]        = get(text, r"Accident Date and Time\s+([\d\-A-Za-z]+)\s*:")
    d["accidentTime"]        = get(text, r"Accident Date and Time\s+[\d\-A-Za-z]+\s*:\s*([\d:]+\s*[APM]+)")
    d["reportingDate"]       = get(text, r"Reporting Date and Time\s+([\d\-A-Za-z]+)\s*:")
    d["reportingTime"]       = get(text, r"Reporting Date and Time\s+[\d\-A-Za-z]+\s*:\s*([\d:]+\s*[APM]+)")
    d["lat"]                 = get(text, r"Lat\s*:\s*([\d.]+)")
    d["lon"]                 = get(text, r"Lon\s*:\s*([\d.]+)")
    d["landmark"]            = get(text, r"Landmark Name\s+(.+?)(?:\n|Location)")
    d["locationDetails"]     = get(text, r"Location Details\s+(.+?)(?:\n|Severity)")
    d["severity"]            = get(text, r"Severity\s+(Fatal|Non-Fatal|Non Fatal)")
    d["numVehicles"]         = get_int(text, r"Number of Vehicle\(s\) involved\s+(\d+)")
    d["roadClassification"]  = get(text, r"Road Classification\s+(.+?)(?:\n|Road Name)")
    d["roadName"]            = get(text, r"Road Name / Street Name\s+(.+?)(?:\n|Local)")
    d["localBody"]           = get(text, r"Local Body\s+(.+?)(?:\n|Accident Spot)")
    d["accidentSpot"]        = get(text, r"Accident Spot\s+(.+?)(?:\n|Collision Type)")
    d["collisionType"]       = get(text, r"Collision Type\s+(.+?)(?:\n|Collision Nature)")
    d["collisionNature"]     = get(text, r"Collision Nature\s+(.+?)(?:\n|Weather)")
    d["weatherCondition"]    = get(text, r"Weather Condition\s+(.+?)(?:\n|Light)")
    d["lightCondition"]      = get(text, r"Light Condition\s+(.+?)(?:\n|Visibility)")
    d["visibility"]          = get(text, r"Visibility\(Approx\.\)\s+(.+?)(?:\n|Initial)")
    d["initialObservation"]  = get(text, r"Initial Observation of Accident Scene\s+(.+?)(?:\n|Traffic)")
    d["trafficViolation"]    = get(text, r"Traffic Violation\s+(.+?)(?:\n|Accident Description)")
    d["accidentDescription"] = get(text, r"Accident Description\s+(.+?)(?:\n|Property)")
    d["propertyDamage"]      = get(text, r"Property Damage\s+(Yes|No)")
    d["approximateDamageValue"] = get(text, r"Approximate Damage Value\s+(.+?)(?:\n|Remedial)")
    d["remedialMeasures"]    = get(text, r"Remedial Measures\s+(.+?)(?:\n|Number of Persons)")

    m = re.search(r"Total\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", text)
    if m:
        d["killed"]         = int(m.group(1))
        d["grievousInjury"] = int(m.group(2))
        d["minorInjury"]    = int(m.group(3))
        d["noInjury"]       = int(m.group(4))
        d["totalPersons"]   = int(m.group(5))
    else:
        d["killed"] = d["grievousInjury"] = d["minorInjury"] = d["noInjury"] = d["totalPersons"] = 0

    d["vehicleRegNumber"]         = get(text, r"Vehicle Registration Number\s+([A-Z0-9]+)")
    d["vehicleOwnerName"]         = get(text, r"Owner Name\s+([A-Z ]+?)(?:\n|Owner Father)")
    d["vehicleOwnerFatherName"]   = get(text, r"Owner Father Name\s+([A-Z ]+?)(?:\n|Owner Address)")
    d["vehicleOwnerAddress"]      = get(text, r"Owner Address\s+(.+?)(?:\n\n|Vehicle Type)")
    d["vehicleType"]              = get(text, r"Vehicle Type\s+(.+?)(?:\n|Load)")
    d["vehicleCategory"]          = get(text, r"Vehicle Category\s+(.+?)(?:\n|Register)")
    d["vehicleColor"]             = get(text, r"Colou?r\s+([A-Z ]+?)(?:\n|Make)")
    d["vehicleMakeModel"]         = get(text, r"Make & Model\s+(.+?)(?:\n|Skid)")
    d["vehicleFuelType"]          = get(text, r"Fuel Type\s+(\w+)")
    d["vehicleYearOfManufacture"] = get(text, r"Year of Manufacture\s+([\d/]+)")
    d["vehicleInsuranceCompany"]  = get(text, r"Insurance Details\s+(.+?)(?:\n|Insurance Policy)")
    d["vehicleInsurancePolicyNumber"] = get(text, r"Insurance Policy Number\s+(\S+)")
    d["vehicleInsuranceValidity"] = get(text, r"Insurance Validity\s+([\d\-A-Za-z]+)")
    d["vehicleFitnessValidity"]   = get(text, r"Fitness Validity\s+([\d\-A-Za-z]+)")
    d["vehiclePucValidity"]       = get(text, r"PUC Certificate Upto\s+([\d\-A-Za-z]+)")
    d["vehicleDamageStatus"]      = get(text, r"Vehicle Damage\s+(.+?)(?:\n|Owner)")
    d["vehicleHitRun"]            = get(text, r"Hit & Run\s+(Yes|No)")

    d["driverName"]          = get(text, r"Driver Name\s+([A-Z ]+?)(?:\n|Guardian)")
    d["driverLicenceNumber"] = get(text, r"Licence Number\s+(\S+)")
    d["driverLicenceType"]   = get(text, r"Driving Licence Type\s+(\w+)")
    d["driverLicenceStatus"] = get(text, r"Driving Licence Status\s+(\w+\.?)")
    d["driverAge"]           = get_int(text, r"Age\s+(\d+)")
    d["driverGender"]        = get(text, r"Gender\s+(Male|Female|Other)")
    d["driverNationality"]   = get(text, r"Nationality\s+(\w+)")
    d["driverBloodGroup"]    = get(text, r"Blood Group\s+([A-Z+\-]+)")
    d["driverMaritalStatus"] = get(text, r"Marital status\s+(\w+)")
    d["driverOccupation"]    = get(text, r"Occupation\s+(\w+)")
    d["driverEducation"]     = get(text, r"Education\s+(\w+)")
    d["driverInjuryType"]    = get(text, r"Injury Type\s+(\w+)")
    d["driverSeverity"]      = get(text, r"Severity\s+(Fatal|Non-Fatal|Grievous|Minor)")
    d["driverSeatbelt"]      = get(text, r"Seatbelt / Helmet\s+(Yes|No)")
    d["driverDrunk"]         = get(text, r"Drunk and Driving\s+(Yes|No)")
    d["driverCellPhone"]     = get(text, r"Cell Phone While Driving\?\s+(Yes|No)")
    d["driverMobile"]        = get(text, r"Current Mobile number\s+(\d+)")
    d["driverAddress"]       = get(text, r"Present Address\s+(.+?)(?:\n\n|Remarks)")
    d["hospitalizationDelay"]   = get(text, r"Hospitalization Delay\s+(.+?)(?:\n|Class)")
    d["modeOfHospitalization"]  = get(text, r"Mode of Hospitalization\s+(.+?)(?:\n|Seatbelt)")

    return d


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Accident Register ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Accident Register"

    DARK_BLUE  = "1A3A6B"
    MED_BLUE   = "2563EB"
    LIGHT_BLUE = "EEF2FF"
    WHITE      = "FFFFFF"

    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color="1A3A6B")
    b_all  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    b_head = Border(left=thick, right=thick, top=thick, bottom=thick)

    def sc(cell, bold=False, size=10, fg=None, bg=None,
           h="left", v="center", wrap=False, border=None):
        cell.font      = Font(name="Arial", bold=bold, size=size, color=fg or "000000")
        if bg:
            cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        if border:
            cell.border = border

    # Row 1 – title
    ws.merge_cells("A1:M1")
    ws["A1"] = (f"Road Accident Register — "
                f"{data.get('stationName','Panvel Town')}, "
                f"{data.get('districtName','Navi Mumbai')} (2025)")
    sc(ws["A1"], bold=True, size=14, fg=WHITE, bg=DARK_BLUE, h="center", border=b_head)
    ws.row_dimensions[1].height = 28

    # Row 2 – source
    ws.merge_cells("A2:M2")
    ws["A2"] = (f"Source: Integrated Road Accident Database (iRAD)  |  "
                f"Station: {data.get('stationName','Panvel Town')}  |  Total Records: 1")
    sc(ws["A2"], size=9, fg="444444", bg="DBEAFE", h="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Rows 4-5 – two-level headers
    single = {
        1:"Sr.\nNo.", 2:"FIR\nNumber", 3:"Date of\nAccident", 4:"Time\n(24 hr)",
        7:"Latitude", 8:"Longitude", 9:"Fatal /\nNon-Fatal",
        10:"Grievous /\nNon-Grievous", 11:"Deaths", 12:"Injured",
        13:"Accident Description\n(2-line summary)"
    }
    for ci, label in single.items():
        col = get_column_letter(ci)
        ws.merge_cells(f"{col}4:{col}5")
        ws[f"{col}4"].value = label
        sc(ws[f"{col}4"], bold=True, size=9, fg=WHITE, bg=DARK_BLUE,
           h="center", wrap=True, border=b_head)

    ws.merge_cells("E4:F4")
    ws["E4"].value = "Vehicle Involved"
    sc(ws["E4"], bold=True, size=9, fg=WHITE, bg=MED_BLUE, h="center", border=b_head)
    ws["E5"].value = "Accused Vehicle"
    sc(ws["E5"], bold=True, size=9, fg=WHITE, bg=MED_BLUE, h="center", wrap=True, border=b_head)
    ws["F5"].value = "Victim Vehicle\n(incl. Pedestrian)"
    sc(ws["F5"], bold=True, size=9, fg=WHITE, bg=MED_BLUE, h="center", wrap=True, border=b_head)

    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 26

    # Row 6 – data
    victim_v = ("-" if "Pedestrian" in (data.get("collisionType") or "")
                else (data.get("vehicleType") or "-"))
    acc_desc = (f"{data.get('accidentDescription','')}\n"
                f"{data.get('collisionType','')}: {data.get('collisionNature','')}")
    row = [
        1, data.get("firNumber","-"), data.get("accidentDate","-"),
        data.get("accidentTime","-"), data.get("vehicleType","-"), victim_v,
        data.get("lat","-"), data.get("lon","-"), data.get("severity","-"),
        "Grievous" if (data.get("grievousInjury") or 0) > 0 else "Non-Grievous",
        data.get("killed",0),
        (data.get("grievousInjury") or 0) + (data.get("minorInjury") or 0),
        acc_desc,
    ]
    for ci, val in enumerate(row, 1):
        col  = get_column_letter(ci)
        cell = ws[f"{col}6"]
        cell.value = val
        sc(cell, size=9, bg=LIGHT_BLUE if ci % 2 == 0 else WHITE,
           h="left" if ci == 13 else "center", wrap=(ci == 13), border=b_all)

    ws.row_dimensions[6].height = 40
    for i, w in enumerate([5,12,13,9,17,17,12,12,13,16,7,7,38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    # ── Sheet 2: Full Details ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Full Details")

    sections = [
        ("CASE IDENTIFIERS", [
            ("FIR Number",             data.get("firNumber")),
            ("FIR Date & Time",        data.get("firDate")),
            ("Station Name",           data.get("stationName")),
            ("District Code",          data.get("districtCode")),
            ("District Name",          data.get("districtName")),
            ("Investigating Officer",  data.get("investigatingOfficer")),
            ("Field Officer",          data.get("fieldOfficer")),
        ]),
        ("ACCIDENT DETAILS", [
            ("Accident Date",          data.get("accidentDate")),
            ("Accident Time",          data.get("accidentTime")),
            ("Reporting Date",         data.get("reportingDate")),
            ("Reporting Time",         data.get("reportingTime")),
            ("Latitude",               data.get("lat")),
            ("Longitude",              data.get("lon")),
            ("Landmark",               data.get("landmark")),
            ("Location Details",       data.get("locationDetails")),
            ("Severity",               data.get("severity")),
            ("Number of Vehicles",     data.get("numVehicles")),
            ("Road Classification",    data.get("roadClassification")),
            ("Road Name",              data.get("roadName")),
            ("Local Body",             data.get("localBody")),
            ("Accident Spot",          data.get("accidentSpot")),
            ("Collision Type",         data.get("collisionType")),
            ("Collision Nature",       data.get("collisionNature")),
            ("Weather Condition",      data.get("weatherCondition")),
            ("Light Condition",        data.get("lightCondition")),
            ("Visibility",             data.get("visibility")),
            ("Initial Observation",    data.get("initialObservation")),
            ("Traffic Violation",      data.get("trafficViolation")),
            ("Accident Description",   data.get("accidentDescription")),
            ("Property Damage",        data.get("propertyDamage")),
            ("Approx. Damage Value",   data.get("approximateDamageValue")),
            ("Remedial Measures",      data.get("remedialMeasures")),
        ]),
        ("CASUALTY SUMMARY", [
            ("Killed",                 data.get("killed", 0)),
            ("Grievous Injury",        data.get("grievousInjury", 0)),
            ("Minor Injury",           data.get("minorInjury", 0)),
            ("No Injury",              data.get("noInjury", 0)),
            ("Total Persons",          data.get("totalPersons", 0)),
        ]),
        ("VEHICLE DETAILS", [
            ("Registration No.",       data.get("vehicleRegNumber")),
            ("Owner Name",             data.get("vehicleOwnerName")),
            ("Owner Father Name",      data.get("vehicleOwnerFatherName")),
            ("Owner Address",          data.get("vehicleOwnerAddress")),
            ("Vehicle Type",           data.get("vehicleType")),
            ("Vehicle Category",       data.get("vehicleCategory")),
            ("Make & Model",           data.get("vehicleMakeModel")),
            ("Color",                  data.get("vehicleColor")),
            ("Fuel Type",              data.get("vehicleFuelType")),
            ("Year of Manufacture",    data.get("vehicleYearOfManufacture")),
            ("Insurance Company",      data.get("vehicleInsuranceCompany")),
            ("Insurance Policy No.",   data.get("vehicleInsurancePolicyNumber")),
            ("Insurance Validity",     data.get("vehicleInsuranceValidity")),
            ("Fitness Validity",       data.get("vehicleFitnessValidity")),
            ("PUC Validity",           data.get("vehiclePucValidity")),
            ("Vehicle Damage",         data.get("vehicleDamageStatus")),
            ("Hit & Run",              data.get("vehicleHitRun")),
        ]),
        ("DRIVER / VICTIM DETAILS", [
            ("Driver Name",            data.get("driverName")),
            ("Licence Number",         data.get("driverLicenceNumber")),
            ("Licence Type",           data.get("driverLicenceType")),
            ("Licence Status",         data.get("driverLicenceStatus")),
            ("Age",                    data.get("driverAge")),
            ("Gender",                 data.get("driverGender")),
            ("Nationality",            data.get("driverNationality")),
            ("Blood Group",            data.get("driverBloodGroup")),
            ("Marital Status",         data.get("driverMaritalStatus")),
            ("Occupation",             data.get("driverOccupation")),
            ("Education",              data.get("driverEducation")),
            ("Injury Type",            data.get("driverInjuryType")),
            ("Severity",               data.get("driverSeverity")),
            ("Seatbelt / Helmet",      data.get("driverSeatbelt")),
            ("Drunk Driving",          data.get("driverDrunk")),
            ("Cell Phone While Driving", data.get("driverCellPhone")),
            ("Mobile Number",          data.get("driverMobile")),
            ("Present Address",        data.get("driverAddress")),
            ("Hospitalization Delay",  data.get("hospitalizationDelay")),
            ("Mode of Hospitalization",data.get("modeOfHospitalization")),
        ]),
        ("LEGAL DETAILS", [
            ("Act",                    data.get("act")),
            ("Sections",               data.get("sections")),
        ]),
    ]

    # Title row for sheet 2
    ws2.merge_cells("A1:B1")
    ws2["A1"] = f"iRAD Full Accident Details — FIR {data.get('firNumber','—')}"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws2["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A1"].border    = b_head
    ws2.row_dimensions[1].height = 24
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 55

    current_row = 2
    ALT = "F8FAFC"

    for section_title, fields in sections:
        # Section header
        ws2.merge_cells(f"A{current_row}:B{current_row}")
        cell = ws2[f"A{current_row}"]
        cell.value = section_title
        cell.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=MED_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border    = b_head
        ws2.row_dimensions[current_row].height = 20
        current_row += 1

        for i, (label, value) in enumerate(fields):
            lc = ws2[f"A{current_row}"]
            vc = ws2[f"B{current_row}"]
            lc.value = label
            vc.value = value if value else "—"
            bg = ALT if i % 2 == 1 else WHITE
            lc.font      = Font(name="Arial", bold=True, size=9)
            lc.fill      = PatternFill("solid", fgColor="EEF2FF")
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            lc.border    = b_all
            vc.font      = Font(name="Arial", size=9)
            vc.fill      = PatternFill("solid", fgColor=bg)
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            vc.border    = b_all
            ws2.row_dimensions[current_row].height = 16
            current_row += 1

        current_row += 1  # blank row between sections

    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  PIPELINE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

STEPS = [
    ("extract", "Extract text from PDF"),
    ("parse",   "Parse accident fields"),
    ("excel",   "Build Excel register"),
    ("done",    "Complete"),
]

def render_pipeline(ph, statuses):
    icons = {"done": "✅", "active": "🔴", "pending": "⬜"}
    lines = []
    for key, label in STEPS:
        s = statuses.get(key, "pending")
        color  = "#10B981" if s=="done" else "#E63946" if s=="active" else "#475569"
        weight = "700" if s == "active" else "400"
        lines.append(
            f'<div class="step-row">'
            f'<span style="color:{color};font-weight:{weight}">{icons[s]} {label}</span>'
            f'</div>'
        )
    ph.markdown("".join(lines), unsafe_allow_html=True)

def add_log(ph, msg):
    t = datetime.now().strftime("%H:%M:%S")
    st.session_state.logs.append(f'<span class="log-time">[{t}]</span> {msg}')
    ph.markdown(
        f'<div class="log-box">{"<br>".join(st.session_state.logs[-12:])}</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  UI LAYOUT
# ══════════════════════════════════════════════════════════════════════════════

col_left, col_right = st.columns([1, 1], gap="large")

with col_left:
    st.markdown('<div class="section-label">📂 Upload iRAD PDF</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("iRAD PDF", type=["pdf"], label_visibility="collapsed")
    if uploaded:
        st.success(f"✅ {uploaded.name}  ({uploaded.size/1024:.1f} KB)")

    run_btn = st.button("⚡ Generate Excel", disabled=not uploaded, use_container_width=True)

    st.markdown("---")
    st.markdown('<div class="section-label">🔄 Pipeline</div>', unsafe_allow_html=True)
    pipeline_ph = st.empty()
    render_pipeline(pipeline_ph, {})

    st.markdown("---")
    st.markdown('<div class="section-label">📋 Log</div>', unsafe_allow_html=True)
    log_ph = st.empty()

with col_right:
    st.markdown('<div class="section-label">📊 Parsed Fields</div>', unsafe_allow_html=True)
    data_ph  = st.empty()
    stats_ph = st.empty()
    st.markdown("---")
    st.markdown('<div class="section-label">⬇️ Download</div>', unsafe_allow_html=True)
    dl_ph = st.empty()


# ══════════════════════════════════════════════════════════════════════════════
#  PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

if run_btn and uploaded:
    st.session_state.logs = []
    statuses = {k: "pending" for k, _ in STEPS}
    statuses["extract"] = "active"
    render_pipeline(pipeline_ph, statuses)

    try:
        # 1 – Extract text
        add_log(log_ph, "Extracting text from PDF...")
        file_bytes = uploaded.read()
        text = extract_text_from_pdf(file_bytes)
        statuses["extract"] = "done"; statuses["parse"] = "active"
        render_pipeline(pipeline_ph, statuses)
        add_log(log_ph, f"Extracted {len(text):,} characters from PDF.")

        # 2 – Parse fields
        add_log(log_ph, "Parsing accident fields...")
        data = parse_irad_pdf(text)
        statuses["parse"] = "done"; statuses["excel"] = "active"
        render_pipeline(pipeline_ph, statuses)
        filled = sum(1 for v in data.values() if v and v != 0)
        add_log(log_ph, f"Parsed {filled} fields. FIR: {data.get('firNumber','—')}, {data.get('accidentDate','—')}")

        # Show parsed data
        with data_ph.container():
            st.json({k: v for k, v in data.items() if v and v != 0}, expanded=False)
        with stats_ph.container():
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Killed",   data.get("killed", 0))
            c2.metric("Grievous", data.get("grievousInjury", 0))
            c3.metric("Minor",    data.get("minorInjury", 0))
            c4.metric("Severity", data.get("severity", "—"))

        # 3 – Excel
        add_log(log_ph, "Building Excel file (2 sheets)...")
        excel_bytes = build_excel(data)
        statuses["excel"] = "done"; statuses["done"] = "done"
        render_pipeline(pipeline_ph, statuses)
        add_log(log_ph, f"✅ Excel ready ({len(excel_bytes)//1024} KB) — 2 sheets generated.")

        # Download
        fir     = (data.get("firNumber") or "accident").replace("/", "-")
        date    = (data.get("accidentDate") or "2025").replace("-", "")
        station = (data.get("stationName") or "Panvel").replace(" ", "_")

        with dl_ph.container():
            st.download_button(
                "📊 Download Excel Register", excel_bytes,
                file_name=f"Accident_Register_{station}_{date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.info("📋 Workbook contains 2 sheets: **Accident Register** (summary) + **Full Details** (all fields)")
            st.success("✅ Excel file generated successfully!")

    except ImportError:
        st.error("❌ PyMuPDF not installed. Run:  py -m pip install pymupdf")
        add_log(log_ph, "ERROR: pymupdf missing.")
    except Exception as exc:
        st.error(f"❌ Error: {exc}")
        add_log(log_ph, f"ERROR: {exc}")
        with st.expander("Traceback"):
            st.code(traceback.format_exc())
