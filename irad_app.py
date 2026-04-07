"""
iRAD PDF → Supabase (PostgreSQL) → Dashboard + Leaflet Map
Material Design 3  |  PyMuPDF + Supabase + Streamlit + Leaflet.js

Requirements (requirements.txt):
    streamlit
    pymupdf
    openpyxl
    pandas
    psycopg2-binary
    sqlalchemy

Setup:
    1. Create a free project at https://supabase.com
    2. Go to Settings → Database → Connection string → URI
    3. In Streamlit Cloud: App Settings → Secrets → add:
          DATABASE_URL = "postgresql://postgres:YOUR_PASSWORD@db.xxxx.supabase.co:5432/postgres"
    4. For local dev: create .streamlit/secrets.toml with same content

Usage:
    streamlit run irad_app.py

Map:  Leaflet.js via CDN — no API key needed.
      - Heatmap overlay of all accidents
      - Color-coded markers (red=fatal, orange=grievous, blue=minor)
      - Red circles mark accident-prone zones (2+ accidents within radius)
      - 5 switchable map styles including satellite
"""

import io
import json
import math
import re
import traceback
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import psycopg2
import psycopg2.extras
import streamlit as st
import streamlit.components.v1 as components
from sqlalchemy import create_engine, text

# ── Config ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="iRAD Dashboard",
    page_icon="🚨",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
#  STYLES — Material Design 3
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Display:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');

:root {
  --primary:        #B3261E;
  --on-primary:     #FFFFFF;
  --primary-cont:   #F9DEDC;
  --on-pri-cont:    #410E0B;
  --secondary:      #775652;
  --sec-cont:       #FFDAD6;
  --surface:        #FFFBFF;
  --surf-var:       #F4DDDB;
  --on-surface:     #201A19;
  --on-surf-var:    #534341;
  --outline:        #857370;
  --outline-var:    #D8C2BE;
  --background:     #FFF8F7;
  --success:        #386A20;
  --success-cont:   #C2F08E;
  --warn:           #7D5700;
  --warn-cont:      #FFDDB3;
  --tonal:          #FAF3F1;
  --e1: 0 1px 2px rgba(0,0,0,.12),0 1px 3px 1px rgba(0,0,0,.08);
  --e2: 0 1px 2px rgba(0,0,0,.12),0 2px 6px 2px rgba(0,0,0,.08);
}

html,body,[class*="css"]{
  font-family:'Google Sans','Roboto',sans-serif;
  background:var(--background);
  color:var(--on-surface);
}
.stApp{background:var(--background);}

.top-bar{
  background:var(--primary);padding:14px 24px;
  margin:-1rem -1rem 0 -1rem;
  display:flex;align-items:center;gap:14px;
}
.top-bar-icon{font-size:26px;}
.top-bar-title{
  font-family:'Google Sans Display',sans-serif;
  font-size:22px;font-weight:500;color:#fff;line-height:1;
}
.top-bar-sub{font-size:12px;color:rgba(255,255,255,.75);margin-top:2px;}

.card{
  background:var(--surface);border-radius:12px;
  box-shadow:var(--e1);padding:20px 22px;margin-bottom:14px;
  border:1px solid var(--outline-var);
}
.card-title{
  font-size:13px;font-weight:700;color:var(--on-surf-var);
  letter-spacing:.5px;text-transform:uppercase;margin-bottom:14px;
  display:flex;align-items:center;gap:8px;
}
.card-title::after{content:'';flex:1;height:1px;background:var(--outline-var);}

.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px;}
.kpi{
  background:var(--surface);border-radius:12px;
  padding:16px 18px;box-shadow:var(--e1);
  border:1px solid var(--outline-var);position:relative;overflow:hidden;
}
.kpi::before{
  content:'';position:absolute;top:0;left:0;right:0;height:4px;
  background:var(--primary);border-radius:4px 4px 0 0;
}
.kpi.green::before{background:var(--success);}
.kpi.warn::before{background:var(--warn);}
.kpi-label{font-size:11px;color:var(--on-surf-var);font-weight:500;
  letter-spacing:.3px;text-transform:uppercase;}
.kpi-value{font-family:'Google Sans Display',sans-serif;font-size:32px;
  font-weight:700;color:var(--on-surface);line-height:1.1;margin-top:4px;}
.kpi-sub{font-size:11px;color:var(--on-surf-var);margin-top:4px;}

.upload-hint{text-align:center;padding:10px 0 4px;
  font-size:13px;color:var(--on-surf-var);}

.stButton>button{
  background:var(--primary)!important;color:#fff!important;
  border:none!important;border-radius:20px!important;
  font-family:'Google Sans',sans-serif!important;font-size:14px!important;
  font-weight:500!important;padding:10px 28px!important;
  box-shadow:var(--e1)!important;transition:all .2s!important;
}
.stButton>button:hover{box-shadow:var(--e2)!important;background:#8B1A14!important;}
.stButton>button:disabled{
  background:rgba(31,27,22,.12)!important;
  color:rgba(31,27,22,.38)!important;box-shadow:none!important;
}
div[data-testid="stFileUploader"]>div{
  border:2px dashed var(--outline)!important;
  border-radius:12px!important;background:var(--surf-var)!important;
}
.stSuccess>div,.stInfo>div,.stWarning>div,.stError>div{border-radius:8px!important;}
hr{border-color:var(--outline-var)!important;margin:6px 0!important;}

.stTabs [data-baseweb="tab-list"]{
  background:var(--surf-var);border-radius:10px;padding:4px;
  border:1px solid var(--outline-var);
}
.stTabs [data-baseweb="tab"]{
  border-radius:8px;font-family:'Google Sans',sans-serif;
  font-size:14px;font-weight:500;color:var(--on-surf-var);
}
.stTabs [aria-selected="true"]{
  background:var(--surface)!important;color:var(--primary)!important;
  box-shadow:var(--e1)!important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE — Supabase / PostgreSQL
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_resource
def get_engine():
    """Create a cached SQLAlchemy engine connected to Supabase."""
    db_url = st.secrets["DATABASE_URL"]
    # psycopg2 needs postgresql:// not postgres://
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    return create_engine(db_url, pool_pre_ping=True)


def get_conn():
    """Return a raw psycopg2 connection (for INSERT / DELETE operations)."""
    db_url = st.secrets["DATABASE_URL"]
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    return psycopg2.connect(db_url)


def init_db():
    """Create the accidents table if it does not exist yet."""
    ddl = """
    CREATE TABLE IF NOT EXISTS accidents (
        id                        SERIAL PRIMARY KEY,
        uploaded_at               TEXT,
        fir_number                TEXT UNIQUE,
        fir_date                  TEXT,
        station_name              TEXT,
        district_name             TEXT,
        district_code             TEXT,
        investigating_officer     TEXT,
        field_officer             TEXT,
        act                       TEXT,
        sections                  TEXT,
        accident_date             TEXT,
        accident_time             TEXT,
        reporting_date            TEXT,
        reporting_time            TEXT,
        lat                       DOUBLE PRECISION,
        lon                       DOUBLE PRECISION,
        landmark                  TEXT,
        location_details          TEXT,
        severity                  TEXT,
        num_vehicles              INTEGER,
        road_classification       TEXT,
        road_name                 TEXT,
        local_body                TEXT,
        accident_spot             TEXT,
        collision_type            TEXT,
        collision_nature          TEXT,
        weather_condition         TEXT,
        light_condition           TEXT,
        visibility                TEXT,
        initial_observation       TEXT,
        traffic_violation         TEXT,
        accident_description      TEXT,
        property_damage           TEXT,
        approximate_damage_value  TEXT,
        remedial_measures         TEXT,
        killed                    INTEGER DEFAULT 0,
        grievous_injury           INTEGER DEFAULT 0,
        minor_injury              INTEGER DEFAULT 0,
        no_injury                 INTEGER DEFAULT 0,
        total_persons             INTEGER DEFAULT 0,
        vehicle_reg_number        TEXT,
        vehicle_owner_name        TEXT,
        vehicle_owner_father      TEXT,
        vehicle_owner_address     TEXT,
        vehicle_type              TEXT,
        vehicle_category          TEXT,
        vehicle_color             TEXT,
        vehicle_make_model        TEXT,
        vehicle_fuel_type         TEXT,
        vehicle_year              TEXT,
        insurance_company         TEXT,
        insurance_policy_number   TEXT,
        insurance_validity        TEXT,
        fitness_validity          TEXT,
        puc_validity              TEXT,
        vehicle_damage_status     TEXT,
        hit_and_run               TEXT,
        driver_name               TEXT,
        driver_licence_number     TEXT,
        driver_licence_type       TEXT,
        driver_licence_status     TEXT,
        driver_age                INTEGER,
        driver_gender             TEXT,
        driver_nationality        TEXT,
        driver_blood_group        TEXT,
        driver_marital_status     TEXT,
        driver_occupation         TEXT,
        driver_education          TEXT,
        driver_injury_type        TEXT,
        driver_severity           TEXT,
        driver_seatbelt           TEXT,
        driver_drunk              TEXT,
        driver_cell_phone         TEXT,
        driver_mobile             TEXT,
        driver_address            TEXT,
        hospitalization_delay     TEXT,
        mode_of_hospitalization   TEXT
    );
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(ddl)
        conn.commit()


# Run once on startup
try:
    init_db()
except Exception as e:
    st.error(f"⚠️ Could not connect to Supabase: {e}\n\nCheck your DATABASE_URL in Streamlit Secrets.")
    st.stop()


def fir_exists(fir_number: str) -> bool:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM accidents WHERE fir_number=%s", (fir_number,))
            return cur.fetchone() is not None


def insert_accident(d: dict) -> bool:
    if not d.get("fir_number") or fir_exists(d["fir_number"]):
        return False

    rows = [
        ("uploaded_at",               datetime.now().isoformat()),
        ("fir_number",                d.get("fir_number")),
        ("fir_date",                  d.get("fir_date")),
        ("station_name",              d.get("station_name")),
        ("district_name",             d.get("district_name")),
        ("district_code",             d.get("district_code")),
        ("investigating_officer",     d.get("investigating_officer")),
        ("field_officer",             d.get("field_officer")),
        ("act",                       d.get("act")),
        ("sections",                  d.get("sections")),
        ("accident_date",             d.get("accident_date")),
        ("accident_time",             d.get("accident_time")),
        ("reporting_date",            d.get("reporting_date")),
        ("reporting_time",            d.get("reporting_time")),
        ("lat",                       _float(d.get("lat"))),
        ("lon",                       _float(d.get("lon"))),
        ("landmark",                  d.get("landmark")),
        ("location_details",          d.get("location_details")),
        ("severity",                  d.get("severity")),
        ("num_vehicles",              d.get("num_vehicles")),
        ("road_classification",       d.get("road_classification")),
        ("road_name",                 d.get("road_name")),
        ("local_body",                d.get("local_body")),
        ("accident_spot",             d.get("accident_spot")),
        ("collision_type",            d.get("collision_type")),
        ("collision_nature",          d.get("collision_nature")),
        ("weather_condition",         d.get("weather_condition")),
        ("light_condition",           d.get("light_condition")),
        ("visibility",                d.get("visibility")),
        ("initial_observation",       d.get("initial_observation")),
        ("traffic_violation",         d.get("traffic_violation")),
        ("accident_description",      d.get("accident_description")),
        ("property_damage",           d.get("property_damage")),
        ("approximate_damage_value",  d.get("approximate_damage_value")),
        ("remedial_measures",         d.get("remedial_measures")),
        ("killed",                    d.get("killed", 0)),
        ("grievous_injury",           d.get("grievous_injury", 0)),
        ("minor_injury",              d.get("minor_injury", 0)),
        ("no_injury",                 d.get("no_injury", 0)),
        ("total_persons",             d.get("total_persons", 0)),
        ("vehicle_reg_number",        d.get("vehicle_reg_number")),
        ("vehicle_owner_name",        d.get("vehicle_owner_name")),
        ("vehicle_owner_father",      d.get("vehicle_owner_father")),
        ("vehicle_owner_address",     d.get("vehicle_owner_address")),
        ("vehicle_type",              d.get("vehicle_type")),
        ("vehicle_category",          d.get("vehicle_category")),
        ("vehicle_color",             d.get("vehicle_color")),
        ("vehicle_make_model",        d.get("vehicle_make_model")),
        ("vehicle_fuel_type",         d.get("vehicle_fuel_type")),
        ("vehicle_year",              d.get("vehicle_year")),
        ("insurance_company",         d.get("insurance_company")),
        ("insurance_policy_number",   d.get("insurance_policy_number")),
        ("insurance_validity",        d.get("insurance_validity")),
        ("fitness_validity",          d.get("fitness_validity")),
        ("puc_validity",              d.get("puc_validity")),
        ("vehicle_damage_status",     d.get("vehicle_damage_status")),
        ("hit_and_run",               d.get("hit_and_run")),
        ("driver_name",               d.get("driver_name")),
        ("driver_licence_number",     d.get("driver_licence_number")),
        ("driver_licence_type",       d.get("driver_licence_type")),
        ("driver_licence_status",     d.get("driver_licence_status")),
        ("driver_age",                d.get("driver_age")),
        ("driver_gender",             d.get("driver_gender")),
        ("driver_nationality",        d.get("driver_nationality")),
        ("driver_blood_group",        d.get("driver_blood_group")),
        ("driver_marital_status",     d.get("driver_marital_status")),
        ("driver_occupation",         d.get("driver_occupation")),
        ("driver_education",          d.get("driver_education")),
        ("driver_injury_type",        d.get("driver_injury_type")),
        ("driver_severity",           d.get("driver_severity")),
        ("driver_seatbelt",           d.get("driver_seatbelt")),
        ("driver_drunk",              d.get("driver_drunk")),
        ("driver_cell_phone",         d.get("driver_cell_phone")),
        ("driver_mobile",             d.get("driver_mobile")),
        ("driver_address",            d.get("driver_address")),
        ("hospitalization_delay",     d.get("hospitalization_delay")),
        ("mode_of_hospitalization",   d.get("mode_of_hospitalization")),
    ]

    cols         = ", ".join(r[0] for r in rows)
    vals         = tuple(r[1] for r in rows)
    placeholders = ", ".join(["%s"] * len(rows))   # PostgreSQL uses %s not ?
    sql          = f"INSERT INTO accidents ({cols}) VALUES ({placeholders})"

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, vals)
        conn.commit()
    return True


def _float(v):
    try:
        return float(v)
    except Exception:
        return None


def load_all() -> pd.DataFrame:
    """Load all records from Supabase into a DataFrame."""
    engine = get_engine()
    with engine.connect() as conn:
        return pd.read_sql(
            text("SELECT * FROM accidents ORDER BY id DESC"),
            conn
        )


def delete_fir(fir_number: str):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM accidents WHERE fir_number=%s", (fir_number,))
        conn.commit()


def delete_all():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM accidents")
        conn.commit()


# ══════════════════════════════════════════════════════════════════════════════
#  MAP HELPERS — Haversine + Cluster detection
# ══════════════════════════════════════════════════════════════════════════════

def haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6_371_000
    p = math.pi / 180
    a = (math.sin((lat2 - lat1) * p / 2) ** 2 +
         math.cos(lat1 * p) * math.cos(lat2 * p) *
         math.sin((lon2 - lon1) * p / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(a))


def find_prone_zones(df_map: pd.DataFrame, radius_m: int = 500):
    pts = df_map[["lat", "lon", "fir_number", "severity",
                  "accident_date", "road_name"]].dropna(subset=["lat", "lon"])
    pts = pts.reset_index(drop=True)
    visited = [False] * len(pts)
    zones   = []

    for i in range(len(pts)):
        if visited[i]:
            continue
        cluster_idx = [i]
        for j in range(i + 1, len(pts)):
            if visited[j]:
                continue
            if haversine_m(pts.at[i, "lat"], pts.at[i, "lon"],
                           pts.at[j, "lat"], pts.at[j, "lon"]) <= radius_m:
                cluster_idx.append(j)

        if len(cluster_idx) >= 2:
            for idx in cluster_idx:
                visited[idx] = True
            cp = pts.iloc[cluster_idx]
            zones.append({
                "lat":   float(cp["lat"].mean()),
                "lon":   float(cp["lon"].mean()),
                "count": len(cluster_idx),
                "firs":  ", ".join(cp["fir_number"].dropna().tolist()),
                "roads": ", ".join(cp["road_name"].dropna().unique().tolist()),
            })
    return zones


# ══════════════════════════════════════════════════════════════════════════════
#  LEAFLET MAP BUILDER
# ══════════════════════════════════════════════════════════════════════════════

TILE_CONFIGS = {
    "OpenStreetMap": {
        "url":     "https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png",
        "attr":    "&copy; OpenStreetMap contributors",
        "maxZoom": 19,
        "sub":     "abc",
    },
    "CartoDB Positron (Light)": {
        "url":     "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
        "attr":    "&copy; OpenStreetMap contributors &copy; CARTO",
        "maxZoom": 19,
        "sub":     "abcd",
    },
    "CartoDB Dark Matter": {
        "url":     "https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png",
        "attr":    "&copy; OpenStreetMap contributors &copy; CARTO",
        "maxZoom": 19,
        "sub":     "abcd",
    },
    "Esri Satellite": {
        "url":     "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        "attr":    "Tiles &copy; Esri &mdash; Source: Esri, Maxar, GeoEye",
        "maxZoom": 18,
        "sub":     "abc",
    },
    "OpenTopoMap": {
        "url":     "https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png",
        "attr":    "Map data: &copy; OpenStreetMap contributors, SRTM",
        "maxZoom": 17,
        "sub":     "abc",
    },
}


def build_leaflet_html(markers_data: list, zones: list, radius_m: int,
                       center_lat: float, center_lon: float, tile_key: str) -> str:
    tile   = TILE_CONFIGS[tile_key]
    m_json = json.dumps(markers_data)
    z_json = json.dumps(zones)

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body, html {{ height:100%; font-family:'Roboto',sans-serif; }}
  #map {{ width:100%; height:100vh; }}
  .stats-bar {{
    position:absolute; top:12px; left:50%; transform:translateX(-50%);
    background:white; border-radius:24px; padding:8px 20px;
    box-shadow:0 3px 14px rgba(0,0,0,.25); font-size:12px; font-weight:600;
    z-index:1000; display:flex; gap:18px; white-space:nowrap;
    border:1px solid #D8C2BE; pointer-events:none;
  }}
  .stat {{ display:flex; align-items:center; gap:6px; }}
  .stat-dot {{ width:10px; height:10px; border-radius:50%; flex-shrink:0; }}
  .legend {{
    background:white; border-radius:10px; padding:12px 16px;
    box-shadow:0 2px 10px rgba(0,0,0,.2); font-size:12px;
    border:1px solid #D8C2BE; min-width:180px;
  }}
  .legend h4 {{ margin:0 0 10px; font-size:13px; color:#B3261E; font-weight:700; }}
  .legend-row {{ display:flex; align-items:center; gap:8px; margin-bottom:7px; color:#333; }}
  .ldot {{ width:13px; height:13px; border-radius:50%; flex-shrink:0; border:1.5px solid rgba(0,0,0,.2); }}
  .lzone {{ width:14px; height:14px; border-radius:50%; background:rgba(179,38,30,.18); border:2px solid #B3261E; flex-shrink:0; }}
  .leaflet-popup-content {{ margin:0 !important; padding:0 !important; }}
  .leaflet-popup-content-wrapper {{ padding:0 !important; border-radius:8px !important; overflow:hidden; box-shadow:0 4px 16px rgba(0,0,0,.2) !important; }}
  .leaflet-popup-tip-container {{ margin-top:-1px; }}
  .popup-wrap {{ font-family:'Roboto',sans-serif; font-size:12px; min-width:230px; max-width:270px; }}
  .popup-head {{ padding:8px 12px; font-weight:700; font-size:13px; color:white; background:#B3261E; }}
  .popup-head.zone {{ background:#7D0000; }}
  .popup-body {{ padding:10px 12px; background:#fff; line-height:1.9; border:1px solid #D8C2BE; border-top:none; }}
  .popup-body b {{ color:#410E0B; }}
  .badge {{ display:inline-block; padding:1px 7px; border-radius:10px; font-size:10px; font-weight:700; margin-left:4px; }}
  .badge-fatal    {{ background:#FFEBEE; color:#B3261E; }}
  .badge-grievous {{ background:#FFF3E0; color:#E65100; }}
  .badge-minor    {{ background:#E3F2FD; color:#1565C0; }}
  .leaflet-control-layers {{ font-family:'Roboto',sans-serif !important; font-size:12px !important; border-radius:10px !important; border:1px solid #D8C2BE !important; box-shadow:0 2px 10px rgba(0,0,0,.15) !important; }}
  .leaflet-control-layers-list {{ padding:4px 6px; }}
</style>
</head>
<body>
<div class="stats-bar">
  <div class="stat"><div class="stat-dot" style="background:#E53935"></div><span id="cFatal">0</span>&nbsp;Fatal</div>
  <div class="stat"><div class="stat-dot" style="background:#FB8C00"></div><span id="cGrievous">0</span>&nbsp;Grievous</div>
  <div class="stat"><div class="stat-dot" style="background:#1E88E5"></div><span id="cMinor">0</span>&nbsp;Minor</div>
  <div class="stat"><div class="stat-dot" style="background:#B3261E; border:1.5px solid #7B0000"></div><span id="cZones">0</span>&nbsp;Prone Zone(s)</div>
</div>
<div id="map"></div>
<script>
(function() {{
  const MARKERS = {m_json};
  const ZONES   = {z_json};
  const RADIUS  = {radius_m};

  const map = L.map('map', {{
    center: [{center_lat}, {center_lon}],
    zoom: 13,
    zoomControl: true,
    preferCanvas: true,
  }});

  L.tileLayer('{tile["url"]}', {{
    attribution: '{tile["attr"]}',
    maxZoom: {tile["maxZoom"]},
    subdomains: '{tile["sub"]}',
  }}).addTo(map);

  let cF=0, cG=0, cM=0;
  MARKERS.forEach(m => {{
    if      (m.color === '#E53935') cF++;
    else if (m.color === '#FB8C00') cG++;
    else                             cM++;
  }});
  document.getElementById('cFatal').textContent    = cF;
  document.getElementById('cGrievous').textContent = cG;
  document.getElementById('cMinor').textContent    = cM;
  document.getElementById('cZones').textContent    = ZONES.length;

  const heatData  = MARKERS.map(m => [m.lat, m.lon, 1]);
  const heatLayer = L.heatLayer(heatData, {{
    radius: 30, blur: 25, maxZoom: 15,
    gradient: {{0.2:'blue', 0.5:'yellow', 0.8:'orange', 1.0:'red'}},
  }});

  function sevBadge(sev) {{
    const s = (sev||'').toLowerCase();
    if (s.includes('fatal') && !s.includes('non'))
      return `<span class="badge badge-fatal">FATAL</span>`;
    if (s.includes('grievous'))
      return `<span class="badge badge-grievous">GRIEVOUS</span>`;
    return `<span class="badge badge-minor">NON-FATAL</span>`;
  }}

  const markerLayer = L.layerGroup();
  MARKERS.forEach(function(pt) {{
    const cm = L.circleMarker([pt.lat, pt.lon], {{
      radius: 9, color: '#7B0000', weight: 2,
      fillColor: pt.color, fillOpacity: 0.95,
    }});
    const popup = `
      <div class="popup-wrap">
        <div class="popup-head">🚨 FIR: ${{pt.fir}}</div>
        <div class="popup-body">
          <div><b>Date / Time:</b> ${{pt.date}} &nbsp;${{pt.time}}</div>
          <div><b>Severity:</b> ${{pt.severity}} ${{sevBadge(pt.severity)}}</div>
          <div><b>Road:</b> ${{pt.road}}</div>
          <div><b>Landmark:</b> ${{pt.landmark}}</div>
          <div><b>Killed:</b> ${{pt.killed}} &nbsp;|&nbsp; <b>Injured:</b> ${{pt.injured}}</div>
          <div><b>Collision:</b> ${{pt.collision}}</div>
          <div><b>Vehicle:</b> ${{pt.vehicle}}</div>
          <div><b>Weather:</b> ${{pt.weather}}</div>
          <div><b>Drunk Driving:</b> ${{pt.drunk}}</div>
          <div style="margin-top:6px;font-size:10px;color:#857370">
            📍 ${{pt.lat.toFixed(6)}}, ${{pt.lon.toFixed(6)}}
          </div>
        </div>
      </div>`;
    cm.bindPopup(popup, {{maxWidth: 290}});
    cm.bindTooltip(`FIR ${{pt.fir}} · ${{pt.severity}}`, {{sticky:true, opacity:0.9}});
    markerLayer.addLayer(cm);
  }});

  const zoneLayer = L.layerGroup();
  ZONES.forEach(function(z) {{
    const zPopup = `
      <div class="popup-wrap">
        <div class="popup-head zone">⚠️ ACCIDENT-PRONE ZONE</div>
        <div class="popup-body">
          <div><b>Accidents in zone:</b> ${{z.count}}</div>
          <div><b>FIR Numbers:</b> ${{z.firs}}</div>
          <div><b>Road(s):</b> ${{z.roads || '—'}}</div>
          <div><b>Zone Radius:</b> ${{RADIUS}} m</div>
          <div style="margin-top:6px;font-size:10px;color:#857370">
            📍 Centre: ${{z.lat.toFixed(6)}}, ${{z.lon.toFixed(6)}}
          </div>
        </div>
      </div>`;
    L.circle([z.lat, z.lon], {{
      radius: RADIUS, color: '#B3261E', weight: 2.5,
      fillColor: '#B3261E', fillOpacity: 0.12,
    }}).bindPopup(zPopup, {{maxWidth:290}})
      .bindTooltip(`⚠️ ${{z.count}} accidents within ${{RADIUS}}m`, {{sticky:true}})
      .addTo(zoneLayer);
    L.circleMarker([z.lat, z.lon], {{
      radius: 8, color: '#7B0000', weight: 2,
      fillColor: '#B3261E', fillOpacity: 0.9,
    }}).bindPopup(zPopup, {{maxWidth:290}})
      .bindTooltip(`⚠️ Prone zone · ${{z.count}} accidents`, {{sticky:true}})
      .addTo(zoneLayer);
  }});

  heatLayer.addTo(map);
  markerLayer.addTo(map);
  zoneLayer.addTo(map);

  const overlayMaps = {{
    '🔥 Accident Heatmap': heatLayer,
    '📍 Accident Markers': markerLayer,
    ['⚠\uFE0F Prone Zones (≤' + RADIUS + 'm)']: zoneLayer,
  }};
  L.control.layers(null, overlayMaps, {{collapsed:false, position:'topright'}}).addTo(map);

  const legend = L.control({{position:'bottomleft'}});
  legend.onAdd = function() {{
    const div = L.DomUtil.create('div', 'legend');
    div.innerHTML = `
      <h4>🚨 iRAD Legend</h4>
      <div class="legend-row"><div class="ldot" style="background:#E53935"></div>Fatal</div>
      <div class="legend-row"><div class="ldot" style="background:#FB8C00"></div>Grievous / Non-Fatal</div>
      <div class="legend-row"><div class="ldot" style="background:#1E88E5"></div>Minor / Other</div>
      <div class="legend-row"><div class="lzone"></div>Prone Zone (≥2 within ${{RADIUS}}m)</div>
    `;
    return div;
  }};
  legend.addTo(map);

  if (MARKERS.length > 1) {{
    const bounds = L.latLngBounds(MARKERS.map(m => [m.lat, m.lon]));
    map.fitBounds(bounds, {{padding:[50, 50]}});
  }} else if (MARKERS.length === 1) {{
    map.setView([MARKERS[0].lat, MARKERS[0].lon], 15);
  }}
}})();
</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  PDF PARSING
# ══════════════════════════════════════════════════════════════════════════════

def extract_text(file_bytes: bytes) -> str:
    import fitz
    doc  = fitz.open(stream=file_bytes, filetype="pdf")
    text = "\n".join(p.get_text() for p in doc)
    doc.close()
    return text


def _get(text, *pats):
    for pat in pats:
        m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
        if m:
            val = m.group(1).strip()
            val = re.split(r'\n[A-Z][a-z]', val)[0].strip()
            return val
    return ""


def _int(text, *pats):
    try:
        return int(_get(text, *pats))
    except Exception:
        return 0


def _extract_fir(text: str) -> str:
    NUM = r'(\d{1,4}[/\\]20\d{2})'
    for pat in [
        r'FIR/CSR\s+Number\s+(?:(?:cr|csr|fir)\s*[Nn]o?\.?\s*[-\u2013]?\s*)?' + NUM,
        r'FIR/CSR\s+Number\s+(?:(?:cr|csr|fir)[-\u2013])?' + NUM,
        r'FIR/CSR\s+Number\s+' + NUM,
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ""


def parse_pdf(text: str) -> dict:
    d = {}
    d["fir_number"]           = _extract_fir(text)
    d["fir_date"]             = _get(text, r"FIR Date & Time\s+([\d\-A-Za-z]+\s*:\s*[\d:]+\s*[APM]+)")
    d["station_name"]         = _get(text, r"Station Name\s+([A-Z][A-Z ]+?)(?:\s{2,}|\n)")
    d["investigating_officer"]= _get(text, r"Investigating(?:g)? Officer\s+([A-Za-z ]+?)(?:\n|Field Officer)")
    d["field_officer"]        = _get(text, r"Field Officer\s+([A-Za-z ]+?)(?:\n|District)")
    d["district_code"]        = _get(text, r"District code\s+(\d+)")
    d["district_name"]        = _get(text, r"District Name\s+([A-Z][A-Z ]+?)(?:\s{2,}|\n|Act)")
    act_block                 = _get(text, r"Act\s+(Bharatiya.*?)\s*Section", r"Act\s+([\w ,\d]+)\n")
    d["act"]                  = act_block.replace("\n", " ").strip()
    d["sections"]             = _get(text, r"Section\s+(.*?)\s*State Rule",
                                     r"(Section[\d\w(),\n ]+?)(?:State Rule|Accident Details)")
    d["accident_date"]        = _get(text, r"Accident Date and Time\s+([\d\-A-Za-z]+)\s*:")
    d["accident_time"]        = _get(text, r"Accident Date and Time\s+[\d\-A-Za-z]+\s*:\s*([\d:]+\s*[APM]+)")
    d["reporting_date"]       = _get(text, r"Reporting Date and Time\s+([\d\-A-Za-z]+)\s*:")
    d["reporting_time"]       = _get(text, r"Reporting Date and Time\s+[\d\-A-Za-z]+\s*:\s*([\d:]+\s*[APM]+)")
    d["lat"]                  = _get(text, r"(?:Geolocation\s+)?Lat\s*:\s*([\d.]+)")
    d["lon"]                  = _get(text, r"Lon\s*:\s*([\d.]+)")
    d["landmark"]             = _get(text, r"Landmark Name\s+(.+?)(?:\n|Location)")
    d["location_details"]     = _get(text, r"Location Details\s+(.+?)(?:\n|Severity)")
    d["severity"]             = _get(text, r"Severity\s+(Fatal|Non-Fatal|Non Fatal)")
    d["num_vehicles"]         = _int(text, r"Number of Vehicle\(s\) involved\s+(\d+)")
    d["road_classification"]  = _get(text, r"Road Classification\s+(.+?)(?:\n|Road Name)")
    d["road_name"]            = _get(text, r"Road Name / Street Name\s+(.+?)(?:\n|Local)")
    d["local_body"]           = _get(text, r"Local Body\s+(.+?)(?:\n|Accident Spot)")
    d["accident_spot"]        = _get(text, r"Accident Spot\s+(.+?)(?:\n|Collision Type)")
    d["collision_type"]       = _get(text, r"Collision Type\s+(.+?)(?:\n|Collision Nature)")
    d["collision_nature"]     = _get(text, r"Collision Nature\s+(.+?)(?:\n|Weather)")
    d["weather_condition"]    = _get(text, r"Weather Condition\s+(.+?)(?:\n|Light)")
    d["light_condition"]      = _get(text, r"Light Condition\s+(.+?)(?:\n|Visibility)")
    d["visibility"]           = _get(text, r"Visibility\(Approx\.\)\s+(.+?)(?:\n|Initial)")
    d["initial_observation"]  = _get(text, r"Initial Observation of Accident Scene\s+(.+?)(?:\n|Traffic)")
    d["traffic_violation"]    = _get(text, r"Traffic Violation\s+(.+?)(?:\n|Accident Description)")
    d["accident_description"] = _get(text, r"Accident Description\s+(.+?)(?:\n|Property)")
    d["property_damage"]      = _get(text, r"Property Damage\s+(Yes|No)")
    d["approximate_damage_value"] = _get(text, r"Approximate Damage Value\s+(.+?)(?:\n|Remedial)")
    d["remedial_measures"]    = _get(text, r"Remedial Measures\s+(.+?)(?:\n|Number of Persons)")
    m = re.search(r"Total\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", text)
    if m:
        d["killed"]          = int(m.group(1))
        d["grievous_injury"] = int(m.group(2))
        d["minor_injury"]    = int(m.group(3))
        d["no_injury"]       = int(m.group(4))
        d["total_persons"]   = int(m.group(5))
    else:
        d["killed"] = d["grievous_injury"] = d["minor_injury"] = d["no_injury"] = d["total_persons"] = 0
    d["vehicle_reg_number"]      = _get(text, r"Vehicle Registration Number\s+([A-Z0-9]+)")
    d["vehicle_owner_name"]      = _get(text, r"Owner Name\s+([A-Z ]+?)(?:\n|Owner Father)")
    d["vehicle_owner_father"]    = _get(text, r"Owner Father Name\s+([A-Z ]+?)(?:\n|Owner Address)")
    d["vehicle_owner_address"]   = _get(text, r"Owner Address\s+(.+?)(?:\n\n|Vehicle Type)")
    d["vehicle_type"]            = _get(text, r"Vehicle Type\s+(.+?)(?:\n|Load)")
    d["vehicle_category"]        = _get(text, r"Vehicle Category\s+(.+?)(?:\n|Register)")
    d["vehicle_color"]           = _get(text, r"Colou?r\s+([A-Z ]+?)(?:\n|Make)")
    d["vehicle_make_model"]      = _get(text, r"Make & Model\s+(.+?)(?:\n|Skid)")
    d["vehicle_fuel_type"]       = _get(text, r"Fuel Type\s+(\w+)")
    d["vehicle_year"]            = _get(text, r"Year of Manufacture\s+([\d/]+)")
    d["insurance_company"]       = _get(text, r"Insurance Details\s+(.+?)(?:\n|Insurance Policy)")
    d["insurance_policy_number"] = _get(text, r"Insurance Policy Number\s+(\S+)")
    d["insurance_validity"]      = _get(text, r"Insurance Validity\s+([\d\-A-Za-z]+)")
    d["fitness_validity"]        = _get(text, r"Fitness Validity\s+([\d\-A-Za-z]+)")
    d["puc_validity"]            = _get(text, r"PUC Certificate Upto\s+([\d\-A-Za-z]+)")
    d["vehicle_damage_status"]   = _get(text, r"Vehicle Damage\s+(.+?)(?:\n|Owner)")
    d["hit_and_run"]             = _get(text, r"Hit & Run\s+(Yes|No)")
    d["driver_name"]             = _get(text, r"Driver Name\s+([A-Z ]+?)(?:\n|Guardian)")
    d["driver_licence_number"]   = _get(text, r"Licence Number\s+(\S+)")
    d["driver_licence_type"]     = _get(text, r"Driving Licence Type\s+(\w+)")
    d["driver_licence_status"]   = _get(text, r"Driving Licence Status\s+(\w+\.?)")
    d["driver_age"]              = _int(text, r"Age\s+(\d+)")
    d["driver_gender"]           = _get(text, r"Gender\s+(Male|Female|Other)")
    d["driver_nationality"]      = _get(text, r"Nationality\s+(\w+)")
    d["driver_blood_group"]      = _get(text, r"Blood Group\s+([A-Z+\-]+)")
    d["driver_marital_status"]   = _get(text, r"Marital status\s+(\w+)")
    d["driver_occupation"]       = _get(text, r"Occupation\s+(\w+)")
    d["driver_education"]        = _get(text, r"Education\s+(\w+)")
    d["driver_injury_type"]      = _get(text, r"Injury Type\s+(\w+)")
    d["driver_severity"]         = _get(text, r"Severity\s+(Fatal|Non-Fatal|Grievous|Minor)")
    d["driver_seatbelt"]         = _get(text, r"Seatbelt / Helmet\s+(Yes|No)")
    d["driver_drunk"]            = _get(text, r"Drunk and Driving\s+(Yes|No)")
    d["driver_cell_phone"]       = _get(text, r"Cell Phone While Driving\?\s+(Yes|No)")
    d["driver_mobile"]           = _get(text, r"Current Mobile number\s+(\d+)")
    d["driver_address"]          = _get(text, r"Present Address\s+(.+?)(?:\n\n|Remarks)")
    d["hospitalization_delay"]   = _get(text, r"Hospitalization Delay\s+(.+?)(?:\n|Class)")
    d["mode_of_hospitalization"] = _get(text, r"Mode of Hospitalization\s+(.+?)(?:\n|Seatbelt)")
    return d


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def df_to_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "iRAD Accident Register"
    thin = Side(style="thin", color="D8C2BE")
    b    = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
        cell.font      = Font(name="Roboto", bold=True, size=9, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="B3261E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = b
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(h)) + 4)
    ws.row_dimensions[1].height = 24
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "FAF3F1" if ri % 2 == 0 else "FFFBFF"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Roboto", size=9)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border    = b
        ws.row_dimensions[ri].height = 15
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  TOP APP BAR
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="top-bar">
  <span class="top-bar-icon">🚨</span>
  <div>
    <div class="top-bar-title">iRAD Accident Dashboard</div>
    <div class="top-bar-sub">Integrated Road Accident Database · Supabase · Leaflet.js</div>
  </div>
</div>
<div style="height:20px"></div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TABS
# ══════════════════════════════════════════════════════════════════════════════

tab_dash, tab_map, tab_upload, tab_manage = st.tabs([
    "📊  Dashboard", "🗺️  Accident Map", "📤  Upload PDF", "🗃️  Manage Records"
])


# ─────────────────────────────────────────────────────────────────────────────
#  TAB 1 — DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
with tab_dash:
    df_all = load_all()

    if df_all.empty:
        st.info("📭 No records yet. Go to **Upload PDF** tab to add accident reports.")
    else:
        st.markdown('<div class="card"><div class="card-title">🔍 Filters</div>', unsafe_allow_html=True)
        fc1, fc2, fc3, fc4, fc5 = st.columns([2, 2, 2, 2, 1])

        fir_opts = ["All"] + sorted(df_all["fir_number"].dropna().unique().tolist())
        sel_fir  = fc1.selectbox("FIR Number", fir_opts, key="f_fir")
        sev_opts = ["All"] + sorted(df_all["severity"].dropna().unique().tolist())
        sel_sev  = fc2.selectbox("Severity", sev_opts, key="f_sev")
        sta_opts = ["All"] + sorted(df_all["station_name"].dropna().unique().tolist())
        sel_sta  = fc3.selectbox("Station", sta_opts, key="f_sta")

        def _try_parse(series):
            for fmt in ["%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d-%B-%Y"]:
                try:
                    p = pd.to_datetime(series, format=fmt, errors="coerce")
                    if p.notna().sum() > 0:
                        return p
                except Exception:
                    pass
            return pd.to_datetime(series, errors="coerce", dayfirst=True)

        dates_raw = _try_parse(df_all["accident_date"]).dropna()
        min_d = dates_raw.min().date() if not dates_raw.empty else date(2020, 1, 1)
        max_d = dates_raw.max().date() if not dates_raw.empty else date.today()
        sel_start = fc4.date_input("Start Date", value=min_d, key="f_start")
        sel_end   = fc5.date_input("End Date",   value=max_d, key="f_end")
        st.markdown('</div>', unsafe_allow_html=True)

        df = df_all.copy()
        df["_date"] = _try_parse(df["accident_date"])
        if sel_fir != "All": df = df[df["fir_number"] == sel_fir]
        if sel_sev != "All": df = df[df["severity"]   == sel_sev]
        if sel_sta != "All": df = df[df["station_name"] == sel_sta]
        if df["_date"].notna().any():
            df = df[(df["_date"].isna()) |
                    ((df["_date"].dt.date >= sel_start) & (df["_date"].dt.date <= sel_end))]

        n              = len(df)
        total_killed   = int(df["killed"].sum())
        total_grievous = int(df["grievous_injury"].sum())
        total_minor    = int(df["minor_injury"].sum())
        fatal_count    = int((df["severity"].str.contains("Fatal", na=False) &
                              ~df["severity"].str.contains("Non", na=False)).sum())

        st.markdown(f"""
        <div class="kpi-grid">
          <div class="kpi">
            <div class="kpi-label">Total Cases</div>
            <div class="kpi-value">{n}</div>
            <div class="kpi-sub">Filtered records</div>
          </div>
          <div class="kpi">
            <div class="kpi-label">Fatal Cases</div>
            <div class="kpi-value" style="color:#B3261E">{fatal_count}</div>
            <div class="kpi-sub">{round(fatal_count/n*100) if n else 0}% of cases</div>
          </div>
          <div class="kpi green">
            <div class="kpi-label">Persons Killed</div>
            <div class="kpi-value" style="color:#386A20">{total_killed}</div>
            <div class="kpi-sub">Deaths recorded</div>
          </div>
          <div class="kpi warn">
            <div class="kpi-label">Injured</div>
            <div class="kpi-value" style="color:#7D5700">{total_grievous + total_minor}</div>
            <div class="kpi-sub">Grievous: {total_grievous} · Minor: {total_minor}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        ch1, ch2, ch3 = st.columns(3)
        with ch1:
            st.markdown('<div class="card"><div class="card-title">📊 Severity Breakdown</div>', unsafe_allow_html=True)
            sv = df["severity"].value_counts().reset_index()
            sv.columns = ["Severity", "Count"]
            if not sv.empty: st.bar_chart(sv.set_index("Severity"), color="#B3261E", height=200)
            else: st.caption("No data")
            st.markdown('</div>', unsafe_allow_html=True)
        with ch2:
            st.markdown('<div class="card"><div class="card-title">🚗 Collision Type</div>', unsafe_allow_html=True)
            ct = df["collision_type"].value_counts().head(6).reset_index()
            ct.columns = ["Type", "Count"]
            if not ct.empty: st.bar_chart(ct.set_index("Type"), color="#775652", height=200)
            else: st.caption("No data")
            st.markdown('</div>', unsafe_allow_html=True)
        with ch3:
            st.markdown('<div class="card"><div class="card-title">🌤 Weather</div>', unsafe_allow_html=True)
            wx = df["weather_condition"].value_counts().head(6).reset_index()
            wx.columns = ["Weather", "Count"]
            if not wx.empty: st.bar_chart(wx.set_index("Weather"), color="#715B2E", height=200)
            else: st.caption("No data")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-title">📈 Accidents Over Time</div>', unsafe_allow_html=True)
        ts = df.dropna(subset=["_date"]).copy()
        if not ts.empty:
            ts["month"] = ts["_date"].dt.to_period("M").astype(str)
            monthly = ts.groupby("month").size().reset_index(name="Cases")
            st.line_chart(monthly.set_index("month"), color="#B3261E", height=180)
        else:
            st.caption("No date data available")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-title">📋 Case Records</div>', unsafe_allow_html=True)
        display_cols = ["fir_number","accident_date","accident_time","station_name",
                        "district_name","severity","killed","grievous_injury","minor_injury",
                        "collision_type","road_name","vehicle_type","driver_name",
                        "driver_drunk","hit_and_run"]
        show_cols = [c for c in display_cols if c in df.columns]
        st.dataframe(df[show_cols].rename(columns=lambda x: x.replace("_"," ").title()),
                     use_container_width=True, height=320, hide_index=True)
        ex1, _ = st.columns([1, 4])
        ex1.download_button("⬇️ Export to Excel", df_to_excel(df[show_cols]),
                            file_name=f"iRAD_Export_{date.today()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  TAB 2 — ACCIDENT MAP
# ─────────────────────────────────────────────────────────────────────────────
with tab_map:
    df_map_all = load_all()
    df_map_all["lat"] = pd.to_numeric(df_map_all["lat"], errors="coerce")
    df_map_all["lon"] = pd.to_numeric(df_map_all["lon"], errors="coerce")

    total_records    = len(df_map_all)
    records_with_gps = len(df_map_all.dropna(subset=["lat","lon"]).query("lat != 0 and lon != 0"))
    records_no_gps   = total_records - records_with_gps

    st.markdown(f"""
    <div class="card">
      <div class="card-title">📡 GPS Coverage</div>
      <div style="display:flex;gap:28px;font-size:13px;flex-wrap:wrap">
        <div>
          <span style="font-size:26px;font-weight:700;color:#386A20">{records_with_gps}</span>
          <span style="color:var(--on-surf-var);margin-left:6px">records with GPS</span>
        </div>
        <div>
          <span style="font-size:26px;font-weight:700;color:#B3261E">{records_no_gps}</span>
          <span style="color:var(--on-surf-var);margin-left:6px">missing GPS</span>
        </div>
        <div>
          <span style="font-size:26px;font-weight:700;color:var(--on-surface)">{total_records}</span>
          <span style="color:var(--on-surf-var);margin-left:6px">total in database</span>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("🔍 GPS Diagnostic — see lat/lon stored per FIR"):
        diag = df_map_all[["fir_number","accident_date","station_name","lat","lon"]].copy()
        diag["GPS Status"] = diag.apply(
            lambda r: "✅ Has GPS" if pd.notna(r["lat"]) and r["lat"] != 0 else "❌ Missing GPS", axis=1
        )
        st.dataframe(diag.rename(columns=lambda x: x.replace("_"," ").title()),
                     use_container_width=True, hide_index=True, height=280)

    if records_with_gps == 0:
        st.warning("⚠️ No records have GPS coordinates. Upload PDFs containing Lat/Lon data to enable the map.")
    else:
        st.markdown('<div class="card"><div class="card-title">🎛️ Map Controls</div>', unsafe_allow_html=True)
        mc1, mc2, mc3, mc4 = st.columns([2, 2, 2, 2])
        sev_filter = mc1.selectbox("Severity",
            ["All"] + sorted(df_map_all["severity"].dropna().unique().tolist()), key="map_sev")
        sta_filter = mc2.selectbox("Station",
            ["All"] + sorted(df_map_all["station_name"].dropna().unique().tolist()), key="map_sta")
        radius_m   = mc3.slider("Prone Zone Radius (m)", 100, 1000, 500, step=50)
        map_style  = mc4.selectbox("Map Style", list(TILE_CONFIGS.keys()))
        st.markdown('</div>', unsafe_allow_html=True)

        df_map = df_map_all.copy()
        if sev_filter != "All": df_map = df_map[df_map["severity"]    == sev_filter]
        if sta_filter != "All": df_map = df_map[df_map["station_name"] == sta_filter]

        df_geo = df_map.dropna(subset=["lat","lon"])
        df_geo = df_geo[(df_geo["lat"] != 0) & (df_geo["lon"] != 0)]

        if df_geo.empty:
            st.warning("No GPS data matches current filters.")
        else:
            markers_data = []
            for _, row in df_geo.iterrows():
                sev = str(row.get("severity", "")).lower()
                color = ("#E53935" if ("fatal" in sev and "non" not in sev)
                         else "#FB8C00" if "grievous" in sev else "#1E88E5")
                markers_data.append({
                    "lat":      float(row["lat"]),
                    "lon":      float(row["lon"]),
                    "color":    color,
                    "fir":      str(row.get("fir_number",  "—")),
                    "date":     str(row.get("accident_date","—")),
                    "time":     str(row.get("accident_time","—")),
                    "severity": str(row.get("severity",    "—")),
                    "road":     str(row.get("road_name",   "—")),
                    "landmark": str(row.get("landmark",    "—")),
                    "killed":   int(row.get("killed", 0) or 0),
                    "injured":  int(row.get("grievous_injury", 0) or 0) + int(row.get("minor_injury", 0) or 0),
                    "collision":str(row.get("collision_type",   "—")),
                    "vehicle":  str(row.get("vehicle_type",     "—")),
                    "drunk":    str(row.get("driver_drunk",     "—")),
                    "weather":  str(row.get("weather_condition","—")),
                })

            zones      = find_prone_zones(df_geo, radius_m)
            center_lat = float(df_geo["lat"].mean())
            center_lon = float(df_geo["lon"].mean())

            html_src = build_leaflet_html(
                markers_data, zones, radius_m, center_lat, center_lon, map_style
            )
            components.html(html_src, height=640, scrolling=False)

            if zones:
                st.markdown('<div class="card"><div class="card-title">⚠️ Accident-Prone Zone Details</div>',
                            unsafe_allow_html=True)
                st.dataframe(
                    pd.DataFrame([{
                        "Zone":        f"Zone {i+1}",
                        "Accidents":   z["count"],
                        "FIR Numbers": z["firs"],
                        "Road(s)":     z["roads"] or "—",
                        "Centre Lat":  round(z["lat"], 6),
                        "Centre Lon":  round(z["lon"], 6),
                    } for i, z in enumerate(zones)]),
                    use_container_width=True, hide_index=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info(f"ℹ️ No prone zones at {radius_m}m. Try increasing the radius.")


# ─────────────────────────────────────────────────────────────────────────────
#  TAB 3 — UPLOAD PDF
# ─────────────────────────────────────────────────────────────────────────────
with tab_upload:
    up_l, up_r = st.columns([1, 1], gap="large")

    with up_l:
        st.markdown('<div class="card"><div class="card-title">📤 Upload iRAD PDF</div>', unsafe_allow_html=True)
        st.markdown('<div class="upload-hint">Drag & drop one or more iRAD PDFs below</div>', unsafe_allow_html=True)
        uploaded_files = st.file_uploader(
            "iRAD PDF files", type=["pdf"], accept_multiple_files=True,
            label_visibility="collapsed",
            key=f"pdf_uploader_{st.session_state.get('uploader_key', 0)}",
        )
        process_btn = st.button("💾  Save to Database", disabled=not uploaded_files, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        total_records = len(load_all())
        st.markdown(f"""
        <div class="card">
          <div class="card-title">🗄️ Database Info</div>
          <p style="font-size:13px;color:var(--on-surf-var);margin:0">
            <b>Engine:</b> Supabase (PostgreSQL)<br>
            <b>Records stored:</b> {total_records}<br>
            <b>Duplicate FIRs:</b> Rejected automatically<br>
            <b>Data persistence:</b> ✅ Permanent (cloud)<br><br>
            <span style="font-size:11px;color:var(--outline)">
            💡 Data is stored in Supabase and survives app restarts.
            </span>
          </p>
        </div>
        """, unsafe_allow_html=True)

    with up_r:
        st.markdown('<div class="card"><div class="card-title">📋 Upload Results</div>', unsafe_allow_html=True)
        results_ph = st.empty()
        results_ph.markdown('<p style="font-size:13px;color:var(--on-surf-var)">Results will appear here after upload.</p>',
                            unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    if process_btn and uploaded_files:
        results = []
        for uf in uploaded_files:
            try:
                raw_bytes = uf.read()
                try:
                    text = extract_text(raw_bytes)
                except ImportError:
                    results.append(("❌", uf.name, "—", "PyMuPDF not installed. Run: pip install pymupdf"))
                    continue

                char_count = len(text)
                with up_r:
                    with st.expander(f"📄 Raw text: {uf.name}  ({char_count:,} chars)"):
                        st.code(text[:3000] if text else "(empty — may be a scanned image PDF)", language=None)

                if char_count < 50:
                    results.append(("⚠️", uf.name, "—",
                                    f"Only {char_count} chars — PDF may be a scanned image (not text-based)."))
                    continue

                parsed   = parse_pdf(text)
                fir      = parsed.get("fir_number") or "UNKNOWN"
                filled   = {k: v for k, v in parsed.items() if v and v != 0}
                unfilled = [k for k, v in parsed.items() if not v or v == 0]

                with up_r:
                    with st.expander(f"🔍 Parsed: FIR {fir} — {len(filled)} fields found, {len(unfilled)} empty"):
                        if filled:   st.json(filled)
                        if unfilled: st.warning("Empty fields: " + ", ".join(unfilled))

                if not parsed.get("fir_number"):
                    results.append(("⚠️", uf.name, fir, "FIR number not found in PDF text."))
                elif fir_exists(fir):
                    results.append(("🔁", uf.name, fir, "Duplicate — FIR already in database. Skipped."))
                else:
                    ok = insert_accident(parsed)
                    results.append(("✅", uf.name, fir, "Saved successfully to Supabase.") if ok
                                   else ("❌", uf.name, fir, "Database insert failed."))

            except Exception as e:
                results.append(("❌", uf.name, "—", f"{type(e).__name__}: {e}"))
                with up_r:
                    with st.expander("🐛 Full Traceback"):
                        st.code(traceback.format_exc())

        rows_html = ""
        for icon, fname, fir, msg in results:
            bg = ("#E8F5E9" if icon == "✅" else "#FFF3E0" if icon == "🔁"
                  else "#FFEBEE" if icon == "❌" else "#FFF8E1")
            rows_html += f"""
            <div style="display:flex;gap:12px;align-items:flex-start;
                        padding:12px 14px;border-radius:8px;margin-bottom:8px;
                        background:{bg};border:1px solid var(--outline-var)">
              <span style="font-size:20px">{icon}</span>
              <div>
                <div style="font-size:13px;font-weight:700;color:var(--on-surface)">{fname}</div>
                <div style="font-size:11px;color:var(--secondary);margin-top:2px">FIR: <b>{fir}</b></div>
                <div style="font-size:12px;color:var(--on-surf-var);margin-top:4px">{msg}</div>
              </div>
            </div>"""
        results_ph.markdown(rows_html, unsafe_allow_html=True)
        if any(r[0] == "✅" for r in results):
            st.session_state["show_success_popup"] = sum(1 for r in results if r[0] == "✅")

# ── Success popup ──────────────────────────────────────────────────────────────
if st.session_state.get("show_success_popup"):
    saved_count = st.session_state["show_success_popup"]

    @st.dialog("✅ Saved Successfully")
    def success_dialog():
        total_now = len(load_all())
        st.markdown(f"""
        <div style="text-align:center;padding:10px 0">
          <div style="font-size:48px;margin-bottom:8px">🎉</div>
          <div style="font-size:18px;font-weight:700;color:var(--success)">
            {saved_count} record(s) saved to Supabase
          </div>
          <div style="font-size:13px;color:var(--on-surf-var);margin-top:6px">
            Total records in database: <b>{total_now}</b>
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("OK", use_container_width=True):
            st.session_state["show_success_popup"] = None
            st.session_state["uploader_key"] = st.session_state.get("uploader_key", 0) + 1
            st.rerun()

    success_dialog()


# ─────────────────────────────────────────────────────────────────────────────
#  TAB 4 — MANAGE RECORDS
# ─────────────────────────────────────────────────────────────────────────────
with tab_manage:
    df_mgmt = load_all()

    st.markdown('<div class="card"><div class="card-title">🗃️ All Records</div>', unsafe_allow_html=True)
    if df_mgmt.empty:
        st.info("No records in the database yet.")
    else:
        mgmt_cols = ["id","fir_number","accident_date","station_name",
                     "district_name","severity","killed","uploaded_at"]
        show_mgmt = [c for c in mgmt_cols if c in df_mgmt.columns]
        st.dataframe(df_mgmt[show_mgmt].rename(columns=lambda x: x.replace("_"," ").title()),
                     use_container_width=True, height=320, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if not df_mgmt.empty:
        st.markdown('<div class="card"><div class="card-title">🗑️ Delete Records</div>', unsafe_allow_html=True)
        del_tab1, del_tab2, del_tab3 = st.tabs(["Delete One", "Delete Multiple", "Delete All"])

        with del_tab1:
            d1c1, d1c2 = st.columns([3, 2])
            with d1c1:
                del_one = st.selectbox("Select FIR to delete",
                                       df_mgmt["fir_number"].dropna().unique().tolist(), key="del_one_fir")
                rec = df_mgmt[df_mgmt["fir_number"] == del_one].iloc[0]
                st.markdown(f"""
                <div style="background:var(--primary-cont);border-radius:8px;
                            padding:10px 14px;font-size:12px;margin-top:8px;
                            border:1px solid var(--outline-var)">
                  <b>Date:</b> {rec.get("accident_date","—")} &nbsp;|&nbsp;
                  <b>Station:</b> {rec.get("station_name","—")} &nbsp;|&nbsp;
                  <b>Severity:</b> {rec.get("severity","—")} &nbsp;|&nbsp;
                  <b>Killed:</b> {rec.get("killed","—")}
                </div>""", unsafe_allow_html=True)
            with d1c2:
                st.markdown("<br>", unsafe_allow_html=True)
                if "confirm_one" not in st.session_state:
                    st.session_state.confirm_one = False
                if not st.session_state.confirm_one:
                    if st.button("🗑️  Delete this FIR", use_container_width=True, key="del_one_btn"):
                        st.session_state.confirm_one = True
                        st.rerun()
                else:
                    st.warning(f"⚠️ Delete **{del_one}**? Cannot be undone.")
                    cc1, cc2 = st.columns(2)
                    if cc1.button("✅ Yes", use_container_width=True, key="del_one_yes"):
                        delete_fir(del_one)
                        st.session_state.confirm_one = False
                        st.success(f"Deleted {del_one}")
                        st.rerun()
                    if cc2.button("❌ Cancel", use_container_width=True, key="del_one_no"):
                        st.session_state.confirm_one = False
                        st.rerun()

        with del_tab2:
            fir_list = df_mgmt["fir_number"].dropna().unique().tolist()
            selected_firs = st.multiselect("Select FIRs to delete", options=fir_list, key="del_multi_firs")
            if selected_firs:
                st.markdown(f"""
                <div style="background:#FFEBEE;border-radius:8px;padding:10px 14px;
                            font-size:12px;border:1px solid #FFCDD2;margin:8px 0">
                  ⚠️ <b>{len(selected_firs)} FIR(s):</b> {", ".join(selected_firs)}
                </div>""", unsafe_allow_html=True)
                if "confirm_multi" not in st.session_state:
                    st.session_state.confirm_multi = False
                mc1, _ = st.columns([2, 3])
                with mc1:
                    if not st.session_state.confirm_multi:
                        if st.button(f"🗑️  Delete {len(selected_firs)} FIR(s)", use_container_width=True, key="del_multi_btn"):
                            st.session_state.confirm_multi = True
                            st.rerun()
                    else:
                        st.warning("Are you sure? Cannot be undone.")
                        my, mn = st.columns(2)
                        if my.button("✅ Yes", use_container_width=True, key="del_multi_yes"):
                            for fir in selected_firs:
                                delete_fir(fir)
                            st.session_state.confirm_multi = False
                            st.success(f"✅ Deleted {len(selected_firs)} record(s)")
                            st.rerun()
                        if mn.button("❌ Cancel", use_container_width=True, key="del_multi_no"):
                            st.session_state.confirm_multi = False
                            st.rerun()
            else:
                st.caption("Select FIR numbers above to enable deletion.")

        with del_tab3:
            total = len(df_mgmt)
            st.markdown(f"""
            <div style="background:#FFEBEE;border-radius:8px;padding:14px 16px;
                        border:1px solid #FFCDD2;margin-bottom:12px">
              <div style="font-size:14px;font-weight:700;color:#B3261E">⚠️ Danger Zone</div>
              <div style="font-size:12px;color:#534341;margin-top:4px">
                Permanently deletes all <b>{total} records</b> from Supabase. Cannot be undone.
              </div>
            </div>""", unsafe_allow_html=True)
            st.download_button("📊 Export backup first (recommended)", df_to_excel(df_mgmt),
                               file_name=f"iRAD_Backup_{date.today()}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            confirm_text = st.text_input("Type  DELETE ALL  to confirm",
                                         key="del_all_confirm", placeholder="DELETE ALL")
            if confirm_text == "DELETE ALL":
                if st.button("🗑️  Delete ALL Records", key="del_all_btn"):
                    delete_all()
                    st.success("✅ All records deleted.")
                    st.rerun()
            else:
                st.button("🗑️  Delete ALL Records", disabled=True, key="del_all_btn_dis")
                if confirm_text:
                    st.error("Type exactly:  DELETE ALL")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-title">⬇️ Full Database Export</div>', unsafe_allow_html=True)
        st.download_button("📊  Download All Records (Excel)", df_to_excel(df_mgmt),
                           file_name=f"iRAD_FullExport_{date.today()}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.markdown('</div>', unsafe_allow_html=True)
